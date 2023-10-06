import openpyxl

#criar uma planilha(book)
book = openpyxl.Workbook()
#como visualizar paginas existentes
print(book.sheetnames)
#como criar uma pagina
book.create_sheet('Frutas')
#como selecionar uma pagina
frutas_page = book['Frutas']
frutas_page.append(['Frutas','Quantidades','Preço','Total'])
frutas_page.append(['Uva','6','2,50','=b2*c2'])
frutas_page.append(['maça','7','2,50','=b3*c3'])
frutas_page.append(['Pera','8','2,50','=b5*c4'])
frutas_page.append(['Mamao','9','2,50','=b5*c5'])
frutas_page.append(['Goiaba','10','2,50','=b6*c6'])
frutas_page.append(['Batata','11','2,50','=b7*c7'])
frutas_page.append(['Abacate','12','2,50','=b8*c8'])
#Salvar a aplanilha
book.save('Planilha de compras.xlsx')